# app.py
# Control Escolar Web (local) - Flask + SQLite + Excel + PDF

from __future__ import annotations

import os
import io
import re
import shutil
import sqlite3
from dataclasses import dataclass
from datetime import datetime, date
from functools import wraps
from typing import Optional, List, Tuple, Dict, Any
from urllib.parse import urlparse  # NEW

from flask import (
    Flask, render_template, request, redirect, url_for, flash,
    send_file, session, jsonify, abort
)
from werkzeug.security import generate_password_hash, check_password_hash

# Excel
from openpyxl import Workbook, load_workbook

# PDF
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

DB_NAME = os.environ.get("CONTROL_ESCOLAR_DB", "control_escolar.db")
BACKUP_DIR = os.environ.get("CONTROL_ESCOLAR_BACKUPS", "backups")
APP_SECRET = os.environ.get("CONTROL_ESCOLAR_SECRET", "change-me-please")

app = Flask(__name__)
app.secret_key = APP_SECRET

# ----------------- MODELOS -----------------

@dataclass
class GroupConfig:
    id: int
    name: str
    periodo_actual: str
    min_aprobacion: float
    total_trabajos: int
    puntos_trabajos: float
    total_tareas: int
    puntos_tareas: float
    teacher_name: str
    label_participacion: str
    label_glosario: str
    label_revision: str
    label_proyecto: str
    label_extra1: str
    label_extra2: str
    label_extra3: str

@dataclass
class Student:
    id: int
    group_id: int
    numero: int
    nombre: str
    trabajos: int
    tareas: int
    examen: float
    proyecto: float
    participacion: float
    glosario: float
    revision: float
    extra1: float
    extra2: float
    extra3: float
    notas: str

# ----------------- DB UTIL -----------------

def db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys=ON;")
    return conn

def init_db():
    os.makedirs(BACKUP_DIR, exist_ok=True)
    conn = db()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS groups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            periodo_actual TEXT DEFAULT '',
            min_aprobacion REAL DEFAULT 6.0,
            total_trabajos INTEGER DEFAULT 0,
            puntos_trabajos REAL DEFAULT 0,
            total_tareas INTEGER DEFAULT 0,
            puntos_tareas REAL DEFAULT 0,
            teacher_name TEXT DEFAULT '',
            label_participacion TEXT DEFAULT 'Participación',
            label_glosario TEXT DEFAULT 'Glosario',
            label_revision TEXT DEFAULT 'Revisión',
            label_proyecto TEXT DEFAULT 'Proyecto',
            label_extra1 TEXT DEFAULT '',
            label_extra2 TEXT DEFAULT '',
            label_extra3 TEXT DEFAULT ''
        );
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            group_id INTEGER NOT NULL,
            numero INTEGER NOT NULL,
            nombre TEXT NOT NULL,
            trabajos INTEGER DEFAULT 0,
            tareas INTEGER DEFAULT 0,
            examen REAL DEFAULT 0,
            proyecto REAL DEFAULT 0,
            participacion REAL DEFAULT 0,
            glosario REAL DEFAULT 0,
            revision REAL DEFAULT 0,
            extra1 REAL DEFAULT 0,
            extra2 REAL DEFAULT 0,
            extra3 REAL DEFAULT 0,
            notas TEXT DEFAULT '',
            FOREIGN KEY(group_id) REFERENCES groups(id) ON DELETE CASCADE
        );
    """)

    # Unique index to prevent duplicate student numbers in the same group
    cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS ux_students_group_numero ON students(group_id, numero);")

    # Settings table for login (simple)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );
    """)

    # Default admin user if not set
    cur.execute("SELECT value FROM settings WHERE key='admin_user'")
    if cur.fetchone() is None:
        cur.execute("INSERT INTO settings(key,value) VALUES('admin_user', ?)", ("admin",))
    cur.execute("SELECT value FROM settings WHERE key='admin_passhash'")
    if cur.fetchone() is None:
        cur.execute("INSERT INTO settings(key,value) VALUES('admin_passhash', ?)",
                    (generate_password_hash("admin"),))

    # Track backups
    cur.execute("SELECT value FROM settings WHERE key='last_auto_backup_date'")
    if cur.fetchone() is None:
        cur.execute("INSERT INTO settings(key,value) VALUES('last_auto_backup_date', ?)", ("",))

    conn.commit()
    conn.close()

def get_setting(key: str, default: str = "") -> str:
    conn = db()
    row = conn.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
    conn.close()
    return row["value"] if row else default

def set_setting(key: str, value: str):
    conn = db()
    conn.execute(
        "INSERT INTO settings(key,value) VALUES(?,?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (key, value)
    )
    conn.commit()
    conn.close()

# ----------------- AUTH -----------------

def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            # full_path incluye querystring; quita el "?" final si no hay query
            nxt = request.full_path
            if nxt.endswith("?"):
                nxt = nxt[:-1]
            return redirect(url_for("login", next=nxt))
        return fn(*args, **kwargs)
    return wrapper

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        admin_user = get_setting("admin_user", "admin")
        passhash = get_setting("admin_passhash", "")
        if username == admin_user and passhash and check_password_hash(passhash, password):
            session["logged_in"] = True
            flash("Sesión iniciada.", "success")
            fallback = url_for("index")
            nxt_raw = request.args.get("next") or fallback
            nxt = safe_next_url(nxt_raw, fallback)
            return redirect(nxt)
        flash("Usuario o contraseña incorrectos.", "error")
    return render_template("login.html", title="Iniciar sesión")

@app.route("/logout")
@login_required
def logout():
    session.clear()
    flash("Sesión cerrada.", "success")
    return redirect(url_for("login"))

@app.route("/settings/password", methods=["GET", "POST"])
@login_required
def settings_password():
    if request.method == "POST":
        current = request.form.get("current") or ""
        new = request.form.get("new") or ""
        confirm = request.form.get("confirm") or ""
        passhash = get_setting("admin_passhash", "")
        if not passhash or not check_password_hash(passhash, current):
            flash("Contraseña actual incorrecta.", "error")
        elif len(new) < 4:
            flash("La nueva contraseña debe tener al menos 4 caracteres.", "error")
        elif new != confirm:
            flash("Confirmación no coincide.", "error")
        else:
            set_setting("admin_passhash", generate_password_hash(new))
            flash("Contraseña actualizada.", "success")
            return redirect(url_for("index"))
    return render_template("settings_password.html", title="Cambiar contraseña")

# ----------------- HELPERS -----------------

def _parse_int(v, default=0) -> int:
    if v is None:
        return default
    if isinstance(v, int):
        return v
    s = str(v).strip()
    if not s:
        return default
    s = s.replace(",", ".")
    try:
        return int(float(s))
    except ValueError:
        return default

def _parse_float(v, default=0.0) -> float:
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return default
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return default

def clamp(n: float, lo: float, hi: float) -> float:
    try:
        n = float(n)
    except Exception:
        return lo
    return max(lo, min(hi, n))

def sanitize_name(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"\s+", " ", s)
    return s

def safe_next_url(next_url: Optional[str], fallback: str) -> str:
    """
    Acepta:
      - rutas locales: "/group/1?page=2"
      - URLs completas del mismo sitio: "http://127.0.0.1:5000/group/1?page=2"
    Rechaza:
      - URLs externas (open redirect)
    """
    if not next_url:
        return fallback

    s = str(next_url).strip()
    if not s:
        return fallback

    p = urlparse(s)

    # Si es URL absoluta (tiene scheme/netloc), convertimos a path+query
    if p.scheme or p.netloc:
        path = p.path or ""
        if not path.startswith("/"):
            return fallback
        if p.query:
            return f"{path}?{p.query}"
        return path

    # Si ya es ruta local
    if not s.startswith("/"):
        return fallback
    return s

# ----------------- DOMAIN -----------------

def get_groups() -> List[GroupConfig]:
    conn = db()
    rows = conn.execute("SELECT * FROM groups ORDER BY name").fetchall()
    conn.close()
    out: List[GroupConfig] = []
    for r in rows:
        out.append(GroupConfig(
            id=r["id"], name=r["name"],
            periodo_actual=r["periodo_actual"] or "",
            min_aprobacion=float(r["min_aprobacion"] or 6.0),
            total_trabajos=int(r["total_trabajos"] or 0),
            puntos_trabajos=float(r["puntos_trabajos"] or 0.0),
            total_tareas=int(r["total_tareas"] or 0),
            puntos_tareas=float(r["puntos_tareas"] or 0.0),
            teacher_name=r["teacher_name"] or "",
            label_participacion=r["label_participacion"] or "Participación",
            label_glosario=r["label_glosario"] or "Glosario",
            label_revision=r["label_revision"] or "Revisión",
            label_proyecto=r["label_proyecto"] or "Proyecto",
            label_extra1=r["label_extra1"] or "",
            label_extra2=r["label_extra2"] or "",
            label_extra3=r["label_extra3"] or "",
        ))
    return out

def get_group(group_id: int) -> Optional[GroupConfig]:
    conn = db()
    r = conn.execute("SELECT * FROM groups WHERE id=?", (group_id,)).fetchone()
    conn.close()
    if not r:
        return None
    return GroupConfig(
        id=r["id"], name=r["name"],
        periodo_actual=r["periodo_actual"] or "",
        min_aprobacion=float(r["min_aprobacion"] or 6.0),
        total_trabajos=int(r["total_trabajos"] or 0),
        puntos_trabajos=float(r["puntos_trabajos"] or 0.0),
        total_tareas=int(r["total_tareas"] or 0),
        puntos_tareas=float(r["puntos_tareas"] or 0.0),
        teacher_name=r["teacher_name"] or "",
        label_participacion=r["label_participacion"] or "Participación",
        label_glosario=r["label_glosario"] or "Glosario",
        label_revision=r["label_revision"] or "Revisión",
        label_proyecto=r["label_proyecto"] or "Proyecto",
        label_extra1=r["label_extra1"] or "",
        label_extra2=r["label_extra2"] or "",
        label_extra3=r["label_extra3"] or "",
    )

def create_group(name: str) -> int:
    name = sanitize_name(name)
    if not name:
        raise ValueError("Nombre de grupo requerido.")
    conn = db()
    cur = conn.cursor()
    cur.execute("INSERT INTO groups(name) VALUES(?)", (name,))
    conn.commit()
    gid = cur.lastrowid
    conn.close()
    return int(gid)

def update_group(cfg: GroupConfig):
    conn = db()
    conn.execute("""
        UPDATE groups SET
          periodo_actual=?, min_aprobacion=?,
          total_trabajos=?, puntos_trabajos=?,
          total_tareas=?, puntos_tareas=?,
          teacher_name=?,
          label_participacion=?, label_glosario=?, label_revision=?, label_proyecto=?,
          label_extra1=?, label_extra2=?, label_extra3=?
        WHERE id=?
    """, (
        cfg.periodo_actual, cfg.min_aprobacion,
        cfg.total_trabajos, cfg.puntos_trabajos,
        cfg.total_tareas, cfg.puntos_tareas,
        cfg.teacher_name,
        cfg.label_participacion, cfg.label_glosario, cfg.label_revision, cfg.label_proyecto,
        cfg.label_extra1, cfg.label_extra2, cfg.label_extra3,
        cfg.id
    ))
    conn.commit()
    conn.close()

def delete_group(group_id: int):
    conn = db()
    conn.execute("DELETE FROM groups WHERE id=?", (group_id,))
    conn.commit()
    conn.close()

def get_students(group_id: int, search: str = "", page: int = 1, per_page: int = 25) -> Tuple[List[Student], int]:
    page = max(1, int(page or 1))
    per_page = int(per_page or 25)
    per_page = 10 if per_page < 10 else 100 if per_page > 100 else per_page
    search = (search or "").strip().lower()

    conn = db()

    where = "WHERE group_id=?"
    params: List[Any] = [group_id]
    if search:
        where += " AND (lower(nombre) LIKE ? OR CAST(numero AS TEXT) LIKE ?)"
        like = f"%{search}%"
        params += [like, like]

    total = conn.execute(f"SELECT COUNT(*) AS c FROM students {where}", params).fetchone()["c"]

    offset = (page - 1) * per_page
    rows = conn.execute(f"""
        SELECT * FROM students {where}
        ORDER BY numero
        LIMIT ? OFFSET ?
    """, params + [per_page, offset]).fetchall()
    conn.close()

    out: List[Student] = []
    for r in rows:
        out.append(Student(
            id=r["id"], group_id=r["group_id"], numero=r["numero"], nombre=r["nombre"],
            trabajos=r["trabajos"] or 0,
            tareas=r["tareas"] or 0,
            examen=r["examen"] or 0.0,
            proyecto=r["proyecto"] or 0.0,
            participacion=r["participacion"] or 0.0,
            glosario=r["glosario"] or 0.0,
            revision=r["revision"] or 0.0,
            extra1=r["extra1"] or 0.0,
            extra2=r["extra2"] or 0.0,
            extra3=r["extra3"] or 0.0,
            notas=r["notas"] or "",
        ))
    return out, int(total)

def get_student(student_id: int) -> Optional[Student]:
    conn = db()
    r = conn.execute("SELECT * FROM students WHERE id=?", (student_id,)).fetchone()
    conn.close()
    if not r:
        return None
    return Student(
        id=r["id"], group_id=r["group_id"], numero=r["numero"], nombre=r["nombre"],
        trabajos=r["trabajos"] or 0,
        tareas=r["tareas"] or 0,
        examen=r["examen"] or 0.0,
        proyecto=r["proyecto"] or 0.0,
        participacion=r["participacion"] or 0.0,
        glosario=r["glosario"] or 0.0,
        revision=r["revision"] or 0.0,
        extra1=r["extra1"] or 0.0,
        extra2=r["extra2"] or 0.0,
        extra3=r["extra3"] or 0.0,
        notas=r["notas"] or "",
    )

def create_student(group_id: int, nombre: str) -> int:
    nombre = sanitize_name(nombre)
    if not nombre:
        raise ValueError("Nombre requerido.")
    conn = db()
    cur = conn.cursor()
    max_no = cur.execute(
        "SELECT COALESCE(MAX(numero), 0) AS m FROM students WHERE group_id=?",
        (group_id,)
    ).fetchone()["m"]
    numero = int(max_no) + 1
    cur.execute(
        "INSERT INTO students(group_id, numero, nombre) VALUES(?,?,?)",
        (group_id, numero, nombre)
    )
    conn.commit()
    sid = cur.lastrowid
    conn.close()
    return int(sid)

def delete_student(student_id: int):
    conn = db()
    conn.execute("DELETE FROM students WHERE id=?", (student_id,))
    conn.commit()
    conn.close()

def validate_student(st: Student, cfg: GroupConfig) -> Student:
    st.nombre = sanitize_name(st.nombre)
    st.trabajos = max(0, int(st.trabajos or 0))
    st.tareas = max(0, int(st.tareas or 0))

    if cfg.total_trabajos > 0 and st.trabajos > cfg.total_trabajos:
        st.trabajos = cfg.total_trabajos
    if cfg.total_tareas > 0 and st.tareas > cfg.total_tareas:
        st.tareas = cfg.total_tareas

    for attr in ("examen", "proyecto", "participacion", "glosario", "revision", "extra1", "extra2", "extra3"):
        setattr(st, attr, clamp(getattr(st, attr) or 0.0, 0.0, 10.0))

    st.notas = (st.notas or "").strip()
    return st

def update_student(st: Student):
    conn = db()
    conn.execute("""
        UPDATE students SET
          nombre=?, trabajos=?, tareas=?,
          examen=?, proyecto=?,
          participacion=?, glosario=?, revision=?,
          extra1=?, extra2=?, extra3=?,
          notas=?
        WHERE id=?
    """, (
        st.nombre, st.trabajos, st.tareas,
        st.examen, st.proyecto,
        st.participacion, st.glosario, st.revision,
        st.extra1, st.extra2, st.extra3,
        st.notas,
        st.id
    ))
    conn.commit()
    conn.close()

def calcular_calificacion_y_criterio(st: Student, cfg: GroupConfig) -> Tuple[float, str]:
    if cfg.total_trabajos > 0:
        pct_trabajos = (st.trabajos / cfg.total_trabajos) * cfg.puntos_trabajos
    else:
        pct_trabajos = 0.0

    if cfg.total_tareas > 0:
        pct_tareas = (st.tareas / cfg.total_tareas) * cfg.puntos_tareas
    else:
        pct_tareas = 0.0

    cal = (
        pct_trabajos + pct_tareas +
        (st.examen or 0) + (st.proyecto or 0) +
        (st.participacion or 0) + (st.glosario or 0) + (st.revision or 0) +
        (st.extra1 or 0) + (st.extra2 or 0) + (st.extra3 or 0)
    )

    if cal <= 0:
        return 0.0, "Pendiente de evaluar"

    criterio = "Aprobado" if cal >= cfg.min_aprobacion else "Reprobado"
    return float(cal), criterio

# ----------------- BACKUPS -----------------

def make_backup(tag: str = "manual") -> str:
    os.makedirs(BACKUP_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = os.path.basename(DB_NAME)
    dst = os.path.join(BACKUP_DIR, f"{base}.{tag}.{ts}.bak")
    shutil.copy2(DB_NAME, dst)
    return dst

def auto_backup_daily():
    today = date.today().isoformat()
    last = get_setting("last_auto_backup_date", "")
    if last == today:
        return
    if os.path.exists(DB_NAME):
        make_backup("auto")
        set_setting("last_auto_backup_date", today)

# ----------------- EXCEL -----------------

EXCEL_HEADERS = [
    "No",
    "Nombre del Alumno",
    "Trabajos entregados",
    "% del total de trabajos entregados",
    "tareas entregadas",
    "% del total de tareas entregados",
    "Examen",
    "Proyecto",
    "Participación",
    "Glosario",
    "Revisión",
    "Calificacion",
    "Criterio",
    "Notas",
]

def excel_template_bytes(cfg: Optional[GroupConfig] = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Grupo"
    headers = EXCEL_HEADERS.copy()
    if cfg:
        headers[7] = cfg.label_proyecto or "Proyecto"
        headers[8] = cfg.label_participacion or "Participación"
        headers[9] = cfg.label_glosario or "Glosario"
        headers[10] = cfg.label_revision or "Revisión"
    ws.append(headers)
    ws.append([1, "EJEMPLO APELLIDO NOMBRE", 0, 0, 0, 0, 0, 0, 0, 0, 0, "", "", ""])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

def export_group_excel(cfg: GroupConfig, students: List[Student]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Grupo"
    headers = EXCEL_HEADERS.copy()
    headers[7] = cfg.label_proyecto or "Proyecto"
    headers[8] = cfg.label_participacion or "Participación"
    headers[9] = cfg.label_glosario or "Glosario"
    headers[10] = cfg.label_revision or "Revisión"
    ws.append(headers)

    for st in students:
        cal, crit = calcular_calificacion_y_criterio(st, cfg)
        pct_trab = (st.trabajos / cfg.total_trabajos) * cfg.puntos_trabajos if cfg.total_trabajos > 0 else 0.0
        pct_tar = (st.tareas / cfg.total_tareas) * cfg.puntos_tareas if cfg.total_tareas > 0 else 0.0
        ws.append([
            st.numero, st.nombre,
            st.trabajos, round(pct_trab, 1),
            st.tareas, round(pct_tar, 1),
            st.examen,
            st.proyecto, st.participacion, st.glosario, st.revision,
            round(cal, 1), crit, st.notas or ""
        ])

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

def parse_excel_preview(excel_bytes: bytes) -> List[Dict[str, Any]]:
    wb = load_workbook(io.BytesIO(excel_bytes), data_only=True)
    ws = wb.active

    rows: List[Dict[str, Any]] = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row:
            continue
        raw_nombre = row[1] if len(row) > 1 else ""
        nombre = sanitize_name(str(raw_nombre or ""))
        if not nombre:
            continue

        no = _parse_int(row[0] if len(row) > 0 else None, default=0)
        trabajos = _parse_int(row[2] if len(row) > 2 else 0)
        tareas = _parse_int(row[4] if len(row) > 4 else 0)
        examen = _parse_float(row[6] if len(row) > 6 else 0)
        proyecto = _parse_float(row[7] if len(row) > 7 else 0)
        participacion = _parse_float(row[8] if len(row) > 8 else 0)
        glosario = _parse_float(row[9] if len(row) > 9 else 0)
        revision = _parse_float(row[10] if len(row) > 10 else 0)
        notas = str(row[13] if len(row) > 13 else "" or "")

        rows.append({
            "rownum": i,
            "numero": no,
            "nombre": nombre,
            "trabajos": trabajos,
            "tareas": tareas,
            "examen": examen,
            "proyecto": proyecto,
            "participacion": participacion,
            "glosario": glosario,
            "revision": revision,
            "notas": notas,
        })
    return rows

def import_excel_apply(cfg: GroupConfig, preview_rows: List[Dict[str, Any]], mode: str):
    """
    mode:
      - replace: delete all students then insert
      - upsert: update by numero, insert if missing; doesn't delete others
    """
    conn = db()
    cur = conn.cursor()

    if mode == "replace":
        cur.execute("DELETE FROM students WHERE group_id=?", (cfg.id,))

    # Determine next number for rows with numero <= 0
    max_no = cur.execute(
        "SELECT COALESCE(MAX(numero),0) AS m FROM students WHERE group_id=?",
        (cfg.id,)
    ).fetchone()["m"]
    next_no = int(max_no) + 1

    for r in preview_rows:
        no = int(r.get("numero") or 0)
        if no <= 0:
            no = next_no
            next_no += 1

        st = Student(
            id=0, group_id=cfg.id, numero=no,
            nombre=r.get("nombre", ""),
            trabajos=_parse_int(r.get("trabajos", 0)),
            tareas=_parse_int(r.get("tareas", 0)),
            examen=_parse_float(r.get("examen", 0.0)),
            proyecto=_parse_float(r.get("proyecto", 0.0)),
            participacion=_parse_float(r.get("participacion", 0.0)),
            glosario=_parse_float(r.get("glosario", 0.0)),
            revision=_parse_float(r.get("revision", 0.0)),
            extra1=0.0, extra2=0.0, extra3=0.0,
            notas=str(r.get("notas", "") or "")
        )
        st = validate_student(st, cfg)

        if mode == "upsert":
            existing = cur.execute(
                "SELECT id FROM students WHERE group_id=? AND numero=?",
                (cfg.id, no)
            ).fetchone()
            if existing:
                cur.execute("""
                    UPDATE students SET
                      nombre=?, trabajos=?, tareas=?,
                      examen=?, proyecto=?,
                      participacion=?, glosario=?, revision=?,
                      notas=?
                    WHERE id=?
                """, (
                    st.nombre, st.trabajos, st.tareas, st.examen, st.proyecto,
                    st.participacion, st.glosario, st.revision, st.notas,
                    existing["id"]
                ))
            else:
                cur.execute("""
                    INSERT INTO students(group_id, numero, nombre, trabajos, tareas, examen, proyecto,
                                         participacion, glosario, revision, extra1, extra2, extra3, notas)
                    VALUES(?,?,?,?,?,?,?,?,?, ?,0,0,0,?)
                """, (
                    cfg.id, no, st.nombre, st.trabajos, st.tareas, st.examen, st.proyecto,
                    st.participacion, st.glosario, st.revision, st.notas
                ))
        else:
            cur.execute("""
                INSERT INTO students(group_id, numero, nombre, trabajos, tareas, examen, proyecto,
                                     participacion, glosario, revision, extra1, extra2, extra3, notas)
                VALUES(?,?,?,?,?,?,?,?,?, ?,0,0,0,?)
            """, (
                cfg.id, no, st.nombre, st.trabajos, st.tareas, st.examen, st.proyecto,
                st.participacion, st.glosario, st.revision, st.notas
            ))

    conn.commit()
    conn.close()

# ----------------- PDF -----------------

def pdf_alumno_bytes(cfg: GroupConfig, st: Student) -> bytes:
    cal, criterio = calcular_calificacion_y_criterio(st, cfg)
    bio = io.BytesIO()
    c = canvas.Canvas(bio, pagesize=A4)
    width, height = A4
    y = height - 50
    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(width / 2, y, "Boleta de evaluación")
    y -= 30

    c.setFont("Helvetica", 10)
    c.drawString(50, y, f"Grupo: {cfg.name}"); y -= 15
    c.drawString(50, y, f"Periodo: {cfg.periodo_actual}"); y -= 15
    if cfg.teacher_name:
        c.drawString(50, y, f"Profesor: {cfg.teacher_name}"); y -= 15
    c.drawString(50, y, f"Número: {st.numero}"); y -= 15
    c.drawString(50, y, f"Nombre: {st.nombre}"); y -= 25

    def line(label, value):
        nonlocal y
        c.drawString(50, y, f"{label}: {value}")
        y -= 15

    line("Trabajos", st.trabajos)
    line("Tareas", st.tareas)
    line("Examen", st.examen)
    line(cfg.label_proyecto or "Proyecto", st.proyecto)
    line(cfg.label_participacion or "Participación", st.participacion)
    line(cfg.label_glosario or "Glosario", st.glosario)
    line(cfg.label_revision or "Revisión", st.revision)
    if cfg.label_extra1 or st.extra1:
        line(cfg.label_extra1 or "Extra 1", st.extra1)
    if cfg.label_extra2 or st.extra2:
        line(cfg.label_extra2 or "Extra 2", st.extra2)
    if cfg.label_extra3 or st.extra3:
        line(cfg.label_extra3 or "Extra 3", st.extra3)

    y -= 10
    line("Calificación final", f"{cal:.1f}")
    line("Criterio", criterio)

    y -= 15
    c.setFont("Helvetica-Bold", 10)
    c.drawString(50, y, "Notas:"); y -= 15
    c.setFont("Helvetica", 10)
    notas = (st.notas or "").strip()
    while notas:
        chunk, notas = notas[:95], notas[95:]
        c.drawString(50, y, chunk)
        y -= 14
        if y < 80:
            c.showPage()
            y = height - 50
            c.setFont("Helvetica", 10)

    c.setFont("Helvetica-Oblique", 8)
    c.drawCentredString(width / 2, 30, "Designed by Omar, con mucho cariño")
    c.showPage()
    c.save()
    return bio.getvalue()

def pdf_grupo_bytes(cfg: GroupConfig, students: List[Student]) -> bytes:
    bio = io.BytesIO()
    c = canvas.Canvas(bio, pagesize=A4)
    width, height = A4

    y = height - 50
    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(width / 2, y, f"Resumen de grupo - {cfg.name}")
    y -= 20
    c.setFont("Helvetica", 10)
    c.drawString(50, y, f"Periodo: {cfg.periodo_actual}"); y -= 15
    if cfg.teacher_name:
        c.drawString(50, y, f"Profesor: {cfg.teacher_name}"); y -= 20
    else:
        y -= 10

    def header():
        nonlocal y
        c.setFont("Helvetica-Bold", 10)
        c.drawString(50, y, "No")
        c.drawString(80, y, "Nombre")
        c.drawString(320, y, "Calif")
        c.drawString(370, y, "Criterio")
        y -= 15
        c.setFont("Helvetica", 9)

    header()
    for st in students:
        cal, crit = calcular_calificacion_y_criterio(st, cfg)
        if y < 80:
            c.setFont("Helvetica-Oblique", 8)
            c.drawCentredString(width / 2, 30, "Designed by Omar, con mucho cariño")
            c.showPage()
            y = height - 50
            header()

        c.drawString(50, y, str(st.numero))
        name = st.nombre or ""
        if len(name) > 35:
            name = name[:32] + "..."
        c.drawString(80, y, name)
        c.drawString(320, y, f"{cal:.1f}")
        c.drawString(370, y, crit)
        y -= 15

    c.setFont("Helvetica-Oblique", 8)
    c.drawCentredString(width / 2, 30, "Designed by Omar, con mucho cariño")
    c.showPage()
    c.save()
    return bio.getvalue()

# ----------------- UI ROUTES -----------------

@app.before_request
def _startup_tasks():
    if not getattr(app, "_db_inited", False):
        init_db()
        app._db_inited = True
    auto_backup_daily()

@app.route("/")
@login_required
def index():
    groups = get_groups()
    return render_template("index.html", title="Inicio", groups=groups)

@app.route("/group/new", methods=["POST"])
@login_required
def group_new():
    name = request.form.get("name") or ""
    try:
        gid = create_group(name)
        flash("Grupo creado.", "success")
        return redirect(url_for("group_detail", group_id=gid))
    except Exception as e:
        flash(str(e), "error")
        return redirect(url_for("index"))

@app.route("/group/<int:group_id>")
@login_required
def group_detail(group_id: int):
    cfg = get_group(group_id)
    if not cfg:
        abort(404)

    search = request.args.get("q", "")
    page = _parse_int(request.args.get("page", 1), 1)
    per_page = _parse_int(request.args.get("per_page", 25), 25)

    students, total = get_students(group_id, search=search, page=page, per_page=per_page)
    rows = [(st, *calcular_calificacion_y_criterio(st, cfg)) for st in students]
    pages = max(1, (total + per_page - 1) // per_page)

    return render_template(
        "group_detail.html",
        title=f"Grupo {cfg.name}",
        cfg=cfg,
        rows=rows,
        q=search,
        page=page,
        per_page=per_page,
        total=total,
        pages=pages
    )

@app.route("/group/<int:group_id>/update", methods=["POST"])
@login_required
def group_update(group_id: int):
    cfg = get_group(group_id)
    if not cfg:
        abort(404)
    try:
        cfg.periodo_actual = (request.form.get("periodo_actual") or "").strip()
        cfg.teacher_name = (request.form.get("teacher_name") or "").strip()
        cfg.min_aprobacion = clamp(_parse_float(request.form.get("min_aprobacion"), 6.0), 0.0, 10.0)
        cfg.total_trabajos = max(0, _parse_int(request.form.get("total_trabajos"), 0))
        cfg.puntos_trabajos = clamp(_parse_float(request.form.get("puntos_trabajos"), 0.0), 0.0, 100.0)
        cfg.total_tareas = max(0, _parse_int(request.form.get("total_tareas"), 0))
        cfg.puntos_tareas = clamp(_parse_float(request.form.get("puntos_tareas"), 0.0), 0.0, 100.0)

        cfg.label_participacion = (request.form.get("label_participacion") or "Participación").strip()
        cfg.label_glosario = (request.form.get("label_glosario") or "Glosario").strip()
        cfg.label_revision = (request.form.get("label_revision") or "Revisión").strip()
        cfg.label_proyecto = (request.form.get("label_proyecto") or "Proyecto").strip()
        cfg.label_extra1 = (request.form.get("label_extra1") or "").strip()
        cfg.label_extra2 = (request.form.get("label_extra2") or "").strip()
        cfg.label_extra3 = (request.form.get("label_extra3") or "").strip()

        update_group(cfg)
        flash("Configuración guardada.", "success")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for("group_detail", group_id=group_id))

@app.route("/group/<int:group_id>/delete", methods=["POST"])
@login_required
def group_delete(group_id: int):
    delete_group(group_id)
    flash("Grupo eliminado.", "success")
    return redirect(url_for("index"))

@app.route("/group/<int:group_id>/student/create", methods=["POST"])
@login_required
def student_create(group_id: int):
    nombre = request.form.get("nombre") or ""
    try:
        sid = create_student(group_id, nombre)
        flash("Alumno agregado.", "success")

        # Regresa al grupo (si venías de ahí), conservando filtros/paginación si el referrer los trae
        fallback = url_for("group_detail", group_id=group_id)
        nxt = safe_next_url(request.referrer, fallback)

        return redirect(url_for("student_edit", student_id=sid, next=nxt))
    except Exception as e:
        flash(str(e), "error")
        return redirect(url_for("group_detail", group_id=group_id))

@app.route("/student/<int:student_id>/edit", methods=["GET", "POST"])
@login_required
def student_edit(student_id: int):
    st = get_student(student_id)
    if not st:
        abort(404)
    cfg = get_group(st.group_id)
    if not cfg:
        abort(404)

    fallback = url_for("group_detail", group_id=st.group_id)
    next_url = safe_next_url(request.args.get("next") or request.form.get("next"), fallback)

    if request.method == "POST":
        try:
            st.nombre = request.form.get("nombre") or st.nombre
            st.trabajos = _parse_int(request.form.get("trabajos"), st.trabajos)
            st.tareas = _parse_int(request.form.get("tareas"), st.tareas)
            st.examen = _parse_float(request.form.get("examen"), st.examen)
            st.proyecto = _parse_float(request.form.get("proyecto"), st.proyecto)
            st.participacion = _parse_float(request.form.get("participacion"), st.participacion)
            st.glosario = _parse_float(request.form.get("glosario"), st.glosario)
            st.revision = _parse_float(request.form.get("revision"), st.revision)
            st.extra1 = _parse_float(request.form.get("extra1"), st.extra1)
            st.extra2 = _parse_float(request.form.get("extra2"), st.extra2)
            st.extra3 = _parse_float(request.form.get("extra3"), st.extra3)
            st.notas = request.form.get("notas") or ""

            st = validate_student(st, cfg)
            update_student(st)
            flash("Alumno guardado.", "success")

            action = request.form.get("_action") or ""
            if action in ("next", "prev"):
                nxt_id = neighbor_student_id(st.group_id, st.numero, direction=action)
                if nxt_id:
                    return redirect(url_for("student_edit", student_id=nxt_id, next=next_url))

            return redirect(url_for("student_edit", student_id=st.id, next=next_url))
        except Exception as e:
            flash(f"Error: {e}", "error")

    cal, crit = calcular_calificacion_y_criterio(st, cfg)
    prev_id = neighbor_student_id(st.group_id, st.numero, "prev")
    next_id = neighbor_student_id(st.group_id, st.numero, "next")

    return render_template(
        "student_edit.html",
        title=f"Alumno {st.numero}",
        st=st, cfg=cfg, cal=cal, crit=crit,
        prev_id=prev_id, next_id=next_id,
        next_url=next_url
    )

def neighbor_student_id(group_id: int, numero: int, direction: str) -> Optional[int]:
    conn = db()
    if direction == "next":
        r = conn.execute(
            "SELECT id FROM students WHERE group_id=? AND numero>? ORDER BY numero LIMIT 1",
            (group_id, numero)
        ).fetchone()
    else:
        r = conn.execute(
            "SELECT id FROM students WHERE group_id=? AND numero<? ORDER BY numero DESC LIMIT 1",
            (group_id, numero)
        ).fetchone()
    conn.close()
    return int(r["id"]) if r else None

@app.route("/student/<int:student_id>/delete", methods=["POST"])
@login_required
def student_delete_route(student_id: int):
    st = get_student(student_id)
    if not st:
        abort(404)
    gid = st.group_id
    delete_student(student_id)
    flash("Alumno eliminado.", "success")
    return redirect(url_for("group_detail", group_id=gid))

# ----------------- QUICK EDIT API (modal) -----------------

@app.route("/api/student/<int:student_id>")
@login_required
def api_student_get(student_id: int):
    st = get_student(student_id)
    if not st:
        return jsonify({"ok": False, "error": "Not found"}), 404
    cfg = get_group(st.group_id)
    cal, crit = calcular_calificacion_y_criterio(st, cfg) if cfg else (0.0, "Pendiente de evaluar")
    return jsonify({
        "ok": True,
        "student": {
            "id": st.id, "group_id": st.group_id, "numero": st.numero, "nombre": st.nombre,
            "trabajos": st.trabajos, "tareas": st.tareas,
            "examen": st.examen, "proyecto": st.proyecto,
            "participacion": st.participacion, "glosario": st.glosario, "revision": st.revision,
            "extra1": st.extra1, "extra2": st.extra2, "extra3": st.extra3,
            "notas": st.notas,
            "cal": round(cal, 1), "crit": crit
        },
        "labels": {
            "proyecto": (cfg.label_proyecto if cfg else "Proyecto"),
            "participacion": (cfg.label_participacion if cfg else "Participación"),
            "glosario": (cfg.label_glosario if cfg else "Glosario"),
            "revision": (cfg.label_revision if cfg else "Revisión"),
            "extra1": (cfg.label_extra1 if cfg and cfg.label_extra1 else "Extra 1"),
            "extra2": (cfg.label_extra2 if cfg and cfg.label_extra2 else "Extra 2"),
            "extra3": (cfg.label_extra3 if cfg and cfg.label_extra3 else "Extra 3"),
        }
    })

@app.route("/api/student/<int:student_id>", methods=["POST"])
@login_required
def api_student_save(student_id: int):
    st = get_student(student_id)
    if not st:
        return jsonify({"ok": False, "error": "Not found"}), 404
    cfg = get_group(st.group_id)
    if not cfg:
        return jsonify({"ok": False, "error": "Group not found"}), 400

    payload = request.get_json(force=True, silent=True) or {}
    try:
        st.nombre = payload.get("nombre", st.nombre)
        st.trabajos = _parse_int(payload.get("trabajos"), st.trabajos)
        st.tareas = _parse_int(payload.get("tareas"), st.tareas)
        st.examen = _parse_float(payload.get("examen"), st.examen)
        st.proyecto = _parse_float(payload.get("proyecto"), st.proyecto)
        st.participacion = _parse_float(payload.get("participacion"), st.participacion)
        st.glosario = _parse_float(payload.get("glosario"), st.glosario)
        st.revision = _parse_float(payload.get("revision"), st.revision)
        st.extra1 = _parse_float(payload.get("extra1"), st.extra1)
        st.extra2 = _parse_float(payload.get("extra2"), st.extra2)
        st.extra3 = _parse_float(payload.get("extra3"), st.extra3)
        st.notas = payload.get("notas", st.notas)

        st = validate_student(st, cfg)
        update_student(st)
        cal, crit = calcular_calificacion_y_criterio(st, cfg)
        return jsonify({"ok": True, "cal": round(cal, 1), "crit": crit})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400

# ----------------- EXCEL ROUTES -----------------

@app.route("/group/<int:group_id>/excel/template")
@login_required
def excel_download_template(group_id: int):
    cfg = get_group(group_id)
    data = excel_template_bytes(cfg)
    fname = f"Plantilla_{cfg.name if cfg else 'Grupo'}.xlsx"
    return send_file(
        io.BytesIO(data),
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/group/<int:group_id>/excel/export")
@login_required
def excel_export(group_id: int):
    cfg = get_group(group_id)
    if not cfg:
        abort(404)
    students, _ = get_students(group_id, page=1, per_page=100000)
    data = export_group_excel(cfg, students)
    fname = f"Grupo_{cfg.name.replace(' ', '_')}.xlsx"
    return send_file(
        io.BytesIO(data),
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/group/<int:group_id>/excel/import", methods=["POST"])
@login_required
def excel_import_upload(group_id: int):
    cfg = get_group(group_id)
    if not cfg:
        abort(404)
    f = request.files.get("file")
    if not f or not f.filename:
        flash("Selecciona un archivo .xlsx.", "error")
        return redirect(url_for("group_detail", group_id=group_id))
    excel_bytes = f.read()
    try:
        preview = parse_excel_preview(excel_bytes)
        if not preview:
            flash("El Excel no contiene filas válidas.", "error")
            return redirect(url_for("group_detail", group_id=group_id))
        session["import_preview"] = preview
        session["import_group_id"] = group_id
        session["import_filename"] = f.filename
        return redirect(url_for("excel_import_preview", group_id=group_id))
    except Exception as e:
        flash(f"Error al leer Excel: {e}", "error")
        return redirect(url_for("group_detail", group_id=group_id))

@app.route("/group/<int:group_id>/excel/import/preview")
@login_required
def excel_import_preview(group_id: int):
    cfg = get_group(group_id)
    if not cfg:
        abort(404)
    preview = session.get("import_preview") or []
    filename = session.get("import_filename") or "archivo.xlsx"
    if session.get("import_group_id") != group_id:
        flash("La previsualización no coincide con el grupo.", "error")
        return redirect(url_for("group_detail", group_id=group_id))
    return render_template(
        "import_preview.html",
        title="Previsualizar importación",
        cfg=cfg, preview=preview, filename=filename
    )

@app.route("/group/<int:group_id>/excel/import/apply", methods=["POST"])
@login_required
def excel_import_apply_route(group_id: int):
    cfg = get_group(group_id)
    if not cfg:
        abort(404)
    preview = session.get("import_preview") or []
    if not preview:
        flash("No hay previsualización cargada.", "error")
        return redirect(url_for("group_detail", group_id=group_id))
    mode = request.form.get("mode") or "replace"
    mode = mode if mode in ("replace", "upsert") else "replace"
    try:
        import_excel_apply(cfg, preview, mode=mode)
        session.pop("import_preview", None)
        session.pop("import_group_id", None)
        session.pop("import_filename", None)
        flash("Importación completada.", "success")
    except Exception as e:
        flash(f"Error importando: {e}", "error")
    return redirect(url_for("group_detail", group_id=group_id))

# ----------------- PDF ROUTES -----------------

@app.route("/student/<int:student_id>/pdf")
@login_required
def pdf_alumno(student_id: int):
    st = get_student(student_id)
    if not st:
        abort(404)
    cfg = get_group(st.group_id)
    if not cfg:
        abort(404)
    data = pdf_alumno_bytes(cfg, st)
    fname = f"Boleta_{cfg.name.replace(' ', '_')}_No{st.numero}.pdf"
    return send_file(io.BytesIO(data), as_attachment=True, download_name=fname, mimetype="application/pdf")

@app.route("/group/<int:group_id>/pdf")
@login_required
def pdf_grupo(group_id: int):
    cfg = get_group(group_id)
    if not cfg:
        abort(404)
    students, _ = get_students(group_id, page=1, per_page=100000)
    data = pdf_grupo_bytes(cfg, students)
    fname = f"Calificaciones_{cfg.name.replace(' ', '_')}.pdf"
    return send_file(io.BytesIO(data), as_attachment=True, download_name=fname, mimetype="application/pdf")

# ----------------- REPORTS -----------------

@app.route("/group/<int:group_id>/report")
@login_required
def group_report(group_id: int):
    cfg = get_group(group_id)
    if not cfg:
        abort(404)
    students, _ = get_students(group_id, page=1, per_page=100000)

    cals = []
    aprobado = 0
    reprobado = 0
    pendiente = 0
    buckets = {"0-5.9": 0, "6-6.9": 0, "7-7.9": 0, "8-8.9": 0, "9-10": 0, "10+": 0}

    for st in students:
        cal, crit = calcular_calificacion_y_criterio(st, cfg)
        cals.append(cal)

        if crit == "Aprobado":
            aprobado += 1
        elif crit == "Reprobado":
            reprobado += 1
        else:
            pendiente += 1

        if cal < 6:
            buckets["0-5.9"] += 1
        elif cal < 7:
            buckets["6-6.9"] += 1
        elif cal < 8:
            buckets["7-7.9"] += 1
        elif cal < 9:
            buckets["8-8.9"] += 1
        elif cal <= 10:
            buckets["9-10"] += 1
        else:
            buckets["10+"] += 1

    n = len(cals)
    avg = round(sum(cals) / n, 2) if n else 0.0
    sorted_cals = sorted(cals)
    median = round(sorted_cals[n // 2], 2) if n else 0.0
    pct_aprob = round((aprobado / n) * 100, 1) if n else 0.0

    scored = []
    for st in students:
        cal, _ = calcular_calificacion_y_criterio(st, cfg)
        scored.append((cal, st))
    scored.sort(key=lambda x: x[0], reverse=True)
    top5 = scored[:5]
    bottom5 = list(reversed(scored[-5:])) if n else []

    return render_template(
        "report_group.html",
        title="Reporte del grupo",
        cfg=cfg,
        n=n, avg=avg, median=median, pct_aprob=pct_aprob,
        aprobado=aprobado, reprobado=reprobado, pendiente=pendiente,
        buckets=buckets,
        top5=top5,
        bottom5=bottom5
    )

# ----------------- BACKUP ROUTES -----------------

@app.route("/backups")
@login_required
def backups_list():
    os.makedirs(BACKUP_DIR, exist_ok=True)
    files = []
    for name in sorted(os.listdir(BACKUP_DIR), reverse=True):
        path = os.path.join(BACKUP_DIR, name)
        if os.path.isfile(path):
            files.append({
                "name": name,
                "size": os.path.getsize(path),
                "mtime": datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m-%d %H:%M:%S")
            })
    return render_template("backups.html", title="Respaldos", files=files)

@app.route("/backups/create", methods=["POST"])
@login_required
def backups_create():
    try:
        dst = make_backup("manual")
        flash(f"Respaldo creado: {os.path.basename(dst)}", "success")
    except Exception as e:
        flash(f"Error creando respaldo: {e}", "error")
    return redirect(url_for("backups_list"))

@app.route("/backups/download/<path:fname>")
@login_required
def backups_download(fname: str):
    path = os.path.join(BACKUP_DIR, fname)
    if not os.path.isfile(path):
        abort(404)
    return send_file(path, as_attachment=True, download_name=fname)

# ----------------- MAIN -----------------

if __name__ == "__main__":
    init_db()
    app.run(host="127.0.0.1", port=int(os.environ.get("PORT", 5000)), debug=True)
